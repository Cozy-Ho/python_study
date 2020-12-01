from openpyxl import Workbook

wb = Workbook()
ws = wb.create_chartsheet()  # 새로운 sheet 생성
ws.title = "MySheet"
# ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2)

new_ws = wb["NewSheet"]  # Dict 형태로 sheet에 접근

new_ws["A1"] = "test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
wb.close()
