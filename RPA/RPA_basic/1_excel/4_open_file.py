from openpyxl import load_workbook  # 파일 불러오기

wb = load_workbook("sample.xlsx")
ws = wb.active

# cell data 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

print()
# cell 갯수를 모를 때
for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
