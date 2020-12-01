from random import *
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.title = "cozy-ho"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["b2"] = 5
ws["b3"] = 6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)  # 없는값은 'none'

print(ws.cell(row=1, column=1).value)  # ws.["A1"].value
print(ws.cell(row=1, column=2).value)  # ws.["B1"].value

c = ws.cell(row=1, column=3, value=10)
print(c.value)

# 반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0, 100))
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")
wb.close()
